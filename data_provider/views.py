# views.py
from django.shortcuts import get_object_or_404
import openpyxl
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Min
from scraper.models import *
from .serializers import *
from rest_framework.permissions import IsAuthenticated
from .pagination import StandardResultsPagination, DashboardResultsPagination
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from django.db.models import Q, Value, CharField, F, Case, When
from django.db.models.functions import Concat
from django.db.models import Prefetch
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from rest_framework.views import APIView
from django.http import Http404, HttpResponse
from django.db.models import OuterRef, Subquery
from calendar import month_name
from django.utils.dateparse import parse_date
from datetime import datetime


class DrugByGenericViewSet(viewsets.ModelViewSet):
    queryset = FSSDrug.objects.all()
    serializer_class = DrugGenericNameSerializer
        
    pagination_class = StandardResultsPagination  # Use the pagination class here

    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "sortOrder",
                openapi.IN_QUERY,
                description="Sort order (ascending or descending)",
                type=openapi.TYPE_STRING,
                enum=["asc", "desc"],
                default="asc",
            ),
            openapi.Parameter(
                "sortColumn",
                openapi.IN_QUERY,
                description="Column to sort by",
                type=openapi.TYPE_STRING,
                default="id",
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "dosage_form",
                openapi.IN_QUERY,
                description="Filter by dosage form",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "route",
                openapi.IN_QUERY,
                description="Filter by route",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def drugs_by_generic_name(self, request):
        try:
            # Fetch filter parameters from request
            sort_order = request.query_params.get("sortOrder", "asc")
            sort_column = request.query_params.get("sortColumn", "id")
            search_text = request.query_params.get("search", "")
            dosage_form = request.query_params.get("dosage_form", "")
            route = request.query_params.get("route", "")

            # Start with all drugs
            queryset = self.get_queryset()

            # Filter by dosage form and route if provided
            if dosage_form:
                queryset = queryset.filter(dosage_form__icontains=dosage_form)
            if route:
                queryset = queryset.filter(route__icontains=route)

            # Search across multiple fields if search is provided
            if search_text:
                queryset = queryset.filter(
                    Q(generic_name__icontains=search_text)
                    | Q(ingredient__icontains=search_text)
                    | Q(strength__icontains=search_text)
                    | Q(dosage_form__icontains=search_text)
                    | Q(route__icontains=search_text)
                )

            # Sorting
            if sort_order == "desc":
                sort_column = f"-{sort_column}"
            queryset = queryset.order_by(sort_column)

            # Group by 'generic_name' and also get the minimum id for each group
            queryset = (
                queryset.values("generic_name")
                .annotate(
                    drug_id=Min("id"),
                    ingredient=Min("ingredient"),
                    strength=Min("strength"),
                    dosage_form=Min("dosage_form"),
                    route=Min("route"),
                )
                .order_by(sort_column)
            )

            # Fetch image_urls for each drug_id after the aggregation
            drug_ids = [drug["drug_id"] for drug in queryset]
            image_urls_map = dict(
                FSSDrug.objects.filter(id__in=drug_ids).values_list("id", "image_urls")
            )

            # Pagination
            page = self.paginate_queryset(queryset)
            if page is not None:

                serialized_data = [
                    {
                        "id": drug["drug_id"],
                        "generic_name": drug["generic_name"],
                        "ingredient": drug["ingredient"],
                        "strength": drug["strength"],
                        "dosage_form": drug["dosage_form"],
                        "route": drug["route"],
                        "image_urls": image_urls_map.get(drug["drug_id"], []),
                    }
                    for drug in page
                ]
                return self.get_paginated_response(serialized_data)

            return Response(
                {
                    "success": True,
                    "message": "No matching records found.",
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error grouping by generic name: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "sortOrder",
                openapi.IN_QUERY,
                description="Sort order (ascending or descending)",
                type=openapi.TYPE_STRING,
                enum=["asc", "desc"],
                default="asc",
            ),
            openapi.Parameter(
                "sortColumn",
                openapi.IN_QUERY,
                description="Column to sort by",
                type=openapi.TYPE_STRING,
                default="id",
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "dosage_form",
                openapi.IN_QUERY,
                description="Filter by dosage form",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "route",
                openapi.IN_QUERY,
                description="Filter by route",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def export_to_excel(self, request):
        try:
            # Fetch filter parameters from request
            sort_order = request.query_params.get("sortOrder", "asc")
            sort_column = request.query_params.get("sortColumn", "id")
            search_text = request.query_params.get("search", "")
            dosage_form = request.query_params.get("dosage_form", "")
            route = request.query_params.get("route", "")

            # Start with all drugs
            queryset = self.get_queryset()

            # Filter by dosage form and route if provided
            if dosage_form:
                queryset = queryset.filter(dosage_form__icontains=dosage_form)
            if route:
                queryset = queryset.filter(route__icontains=route)

            # Search across multiple fields if search is provided
            if search_text:
                queryset = queryset.filter(
                    Q(generic_name__icontains=search_text)
                    | Q(ingredient__icontains=search_text)
                    | Q(strength__icontains=search_text)
                    | Q(dosage_form__icontains=search_text)
                    | Q(route__icontains=search_text)
                )

            # Sorting
            if sort_order == "desc":
                sort_column = f"-{sort_column}"
            queryset = queryset.order_by(sort_column)

            # Group by 'generic_name' and also get the minimum id for each group
            queryset = (
                queryset.values("generic_name")
                .annotate(
                    drug_id=Min("id"),
                    ingredient=Min("ingredient"),
                    strength=Min("strength"),
                    dosage_form=Min("dosage_form"),
                    route=Min("route"),
                )
                .order_by(sort_column)
            )

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Drugs by Generic Name"

            # Add headers to the worksheet
            headers = ["Generic Name", "Ingredient", "Strength", "Dosage Form", "Route"]
            sheet.append(headers)

            # Write data to the worksheet
            for drug in queryset:
                sheet.append(
                    [
                        drug["generic_name"],
                        drug["ingredient"],
                        drug["strength"],
                        drug["dosage_form"],
                        drug["route"],
                    ]
                )

            # Set the file name
            file_name = f"drugs_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DrugByTradeViewSet(viewsets.ModelViewSet):
    queryset = FSSDrug.objects.all()
    serializer_class = DrugTradeNameSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsPagination
    ordering = ["id"]

    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "sortOrder",
                openapi.IN_QUERY,
                description="Sort order (ascending or descending)",
                type=openapi.TYPE_STRING,
                enum=["asc", "desc"],
                default="asc",
            ),
            openapi.Parameter(
                "sortColumn",
                openapi.IN_QUERY,
                description="Column to sort by",
                type=openapi.TYPE_STRING,
                default="id",
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "price_type",
                openapi.IN_QUERY,
                description="Filter by price type (FSS, NC, Big4)",
                type=openapi.TYPE_STRING,
                enum=["FSS", "NC", "Big4"],
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def drug_by_trade_name(self, request):
        try:
            sort_order = request.query_params.get("sortOrder", "asc")
            sort_column = request.query_params.get("sortColumn", "id")
            search_text = request.query_params.get("search", "")
            price_type = request.query_params.get("price_type")

            queryset = self.get_queryset()

            # Apply search filter if provided
            if search_text:
                queryset = queryset.filter(
                    Q(trade_name__icontains=search_text)
                    | Q(ingredient__icontains=search_text)
                    | Q(package_description__icontains=search_text)
                    | Q(ndc_with_dashes__icontains=search_text)
                )

            # Apply price type filter if provided
            if price_type:
                queryset = queryset.filter(pricings__price_type=price_type)

            # Sorting
            if sort_order == "desc":
                sort_column = f"-{sort_column}"
            queryset = queryset.order_by(sort_column)

            # Group by 'trade_name' and get the first entry for each group
            queryset_drugs = (
                queryset.annotate(
                    min_id=Min('id'),
                    min_ingredient=Min('ingredient'),
                    min_strength=Min('strength'),
                    min_ndc_code=Min('ndc_with_dashes'),
                    min_package_description=Min('package_description'),
                    min_contract_stop_date=Min('contract__contract_stop_date'),
                    min_price=Min('pricings__price'),
                    min_price_type=Min('pricings__price_type')
                )
                .values(
                    'min_id', 'trade_name', 'min_strength', 'min_ingredient', 
                    'min_ndc_code', 'min_package_description', 
                    'min_price', 'min_price_type', 'min_contract_stop_date'
                )
                .order_by(sort_column)
            )

            # Apply pagination
            page = self.paginate_queryset(queryset_drugs)
            if page is not None:
                serialized_data = [
                    {
                        "id": drug["min_id"],
                        "trade_name": drug["trade_name"],
                        "strength": drug["min_strength"],
                        "ingredient": drug["min_ingredient"],
                        "ndc_code": drug["min_ndc_code"],
                        "package_description": drug["min_package_description"],
                        "price": drug["min_price"],
                        "price_type": drug["min_price_type"],
                        "contract_stop_date": drug["min_contract_stop_date"],
                    }
                    for drug in page
                ]

                return self.get_paginated_response(serialized_data)

            return Response(
                {
                    "success": True,
                    "message": "Data retrieved successfully",
                    "data": [],
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(f"Error filtering by trade name and price type: {str(e)}")
            return Response(
                {
                    "success": False,
                    "message": f"Error filtering by trade name and price type: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "sortOrder",
                openapi.IN_QUERY,
                description="Sort order (ascending or descending)",
                type=openapi.TYPE_STRING,
                enum=["asc", "desc"],
                default="asc",
            ),
            openapi.Parameter(
                "sortColumn",
                openapi.IN_QUERY,
                description="Column to sort by",
                type=openapi.TYPE_STRING,
                default="id",
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "price_type",
                openapi.IN_QUERY,
                description="Filter by price type (FSS, NC, Big4)",
                type=openapi.TYPE_STRING,
                enum=["FSS", "NC", "Big4"],
                collectionFormat="multi",
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def export_to_excel(self, request):
        try:
            # Fetch filter parameters from request
            sort_order = request.query_params.get("sortOrder", "asc")
            sort_column = request.query_params.get("sortColumn", "id")
            search_text = request.query_params.get("search", "")
            price_type = request.query_params.get("price_type", "")

            # Start with all drugs
            queryset = self.get_queryset()

            # Apply search filter if provided
            if search_text:
                queryset = queryset.filter(
                    Q(trade_name__icontains=search_text)
                    | Q(ingredient__icontains=search_text)
                    | Q(package_description__icontains=search_text)
                    | Q(ndc_with_dashes__icontains=search_text)
                )

            # Filter by price type if provided
            if price_type:
                queryset = queryset.filter(pricings__price_type__iexact=price_type)
            else:
                queryset = queryset.distinct()

            # Sorting
            if sort_order == "desc":
                sort_column = f"-{sort_column}"
            queryset = queryset.order_by(sort_column)

            # Group by 'trade_name' and also get the minimum id for each group
            queryset_drugs = (
                queryset.values("trade_name")
                .annotate(
                    drug_id=Min("id"),
                    ingredient=Min("ingredient"),
                    strength=Min("strength"),
                    ndc_code=Min("ndc_with_dashes"),
                    package_description=Min("package_description"),
                    contract_stop_date=Min(
                        "contract__contract_stop_date"
                    ),  # Add this line
                )
                .order_by(sort_column)
            )

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Drugs by Trade Name"

            # Add headers to the worksheet
            headers = [
                "Trade Name",
                "Ingredient",
                "Strength",
                "NDC Code",
                "Package Description",
                "Price",
                "Price Type",
                "Contract Stop Date",
            ]
            sheet.append(headers)

            # Efficiently fetch prices for all drugs in the queryset in one query
            trade_names = [drug["trade_name"] for drug in queryset_drugs]
            if price_type:
                prices = FSSPricing.objects.filter(
                    drug__trade_name__in=trade_names, price_type__iexact=price_type
                ).values("drug__trade_name", "price", "price_type")
            else:
                prices = FSSPricing.objects.filter(
                    drug__trade_name__in=trade_names
                ).values("drug__trade_name", "price", "price_type")

            price_map = {}
            for price in prices:
                trade_name = price["drug__trade_name"]
                if trade_name not in price_map:
                    price_map[trade_name] = {
                        "price": price["price"],
                        "price_type": price["price_type"],
                    }

            # Write data to the worksheet
            for drug in queryset_drugs:
                sheet.append(
                    [
                        drug["trade_name"],
                        drug["ingredient"],
                        drug["strength"],
                        drug["ndc_code"],
                        drug["package_description"],
                        price_map.get(drug["trade_name"], {}).get("price", None),
                        price_map.get(drug["trade_name"], {}).get("price_type", None),
                        drug.get("contract_stop_date", None),  # Add this line
                    ]
                )

            # Set the file name
            file_name = (
                f"trade_drugs_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class DrugContractorsViewSet(viewsets.ModelViewSet):
    queryset = FSSDrug.objects.all()
    serializer_class = DrugContractorsSerializer
    permission_classes = [IsAuthenticated]
    @swagger_auto_schema(
        method='get',
        manual_parameters=[
            openapi.Parameter(
                'search',
                openapi.IN_QUERY,
                description='Search text for filtering across vendor name, trade name, and contract number',
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                'year',
                openapi.IN_QUERY,
                description='Filter by the year of the contract (can be start or end year)',
                type=openapi.TYPE_INTEGER,
            ),
            openapi.Parameter(
                'month',
                openapi.IN_QUERY,
                description='Filter by the month of the contract (e.g., "March", "April") (can be start or end month)',
                type=openapi.TYPE_STRING,
            ),
        ]
    )
    @action(detail=True, methods=["get"])
    def drug_contractors(self, request, drug_id=None):
        queryset = self.get_queryset()
        # Filter by drug_id provided in the URL
        queryset = queryset.filter(id=drug_id)
        # Additional custom filtering based on query parameters
        search_text = request.query_params.get('search', None)
        year = request.query_params.get('year', None)
        month = request.query_params.get('month', None)
        # Filter by search text
        if search_text:
            queryset = queryset.filter(
                Q(contract__vendor__vendor_name__icontains=search_text) |
                Q(contract__contract_number__icontains=search_text) |
                Q(trade_name__icontains=search_text)
            )
        # Initialize Q object for date filters
        date_filters = Q()
        # Handle year and month filtering
        if year:
            try:
                year = int(year)
                # Apply year filtering on both start and end date
                date_filters &= (Q(contract__contract_start_date__year=year) |
                                 Q(contract__contract_stop_date__year=year))
            except ValueError:
                return Response(
                    {
                        "success": False,
                        "message": "Invalid year format. Please use a valid integer year.",
                        "data": None,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        if month:
            try:
                month = month.capitalize()
                month_number = {month: index for index, month in enumerate(month_name) if month}
                if month not in month_number:
                    raise ValueError("Invalid month name. Please provide a valid month name (e.g., 'March').")
                # Apply month filtering on both start and end date
                date_filters &= (Q(contract__contract_start_date__month=month_number[month]) |
                                 Q(contract__contract_stop_date__month=month_number[month]))
            except ValueError as e:
                return Response(
                    {
                        "success": False,
                        "message": str(e),
                        "data": None,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        # Apply date filters to queryset
        if date_filters:
            queryset = queryset.filter(date_filters)
        # Handle no results found
        if not queryset.exists():
            return Response(
                {
                    "success": False,
                    "message": "No drug contractors found with the provided filters.",
                    "data": [],
                },
                status=status.HTTP_200_OK,
            )
        # Prepare the detailed response data
        response_data = []
        for drug in queryset:
            response_data.append({
                "Vendor_id": drug.vendor.id if drug.vendor else None,
                "vendor_name": drug.vendor.vendor_name if drug.vendor else None,
                "contract_number": drug.contract.contract_number if drug.contract else None,
                "trade_name": drug.trade_name,
                "contract_start_date": drug.contract.contract_start_date.strftime("%Y-%m-%d") if drug.contract else None,
                "contract_stop_date": drug.contract.contract_stop_date.strftime("%Y-%m-%d") if drug.contract else None,
            })
        return Response(
            {
                "success": True,
                "message": "Filtered drug contractors retrieved successfully.",
                "data": response_data,
            },
            status=status.HTTP_200_OK,
        )
# class DrugContractorsViewSet(viewsets.ModelViewSet):
#     queryset = FSSDrug.objects.all()
#     serializer_class = DrugContractorsSerializer
#     permission_classes = [IsAuthenticated]

#     @action(detail=True, methods=["get"])
#     def drug_contractors(self, request, drug_id=None):

#         try:
#             # Fetch the drug by its primary key (ID)
#             drug = get_object_or_404(FSSDrug, id=drug_id)

#             # Prepare response data
#             response_data = {
#                 "id": drug.vendor.id,
#                 "vendor_name": drug.vendor.vendor_name if drug.vendor else None,
#                 "contract_number": drug.contract.contract_number,
#                 "trade_name": drug.trade_name,
#                 "contract_start_date": drug.contract.contract_start_date.strftime(
#                     "%Y-%m-%d"
#                 ),
#                 "contract_stop_date": drug.contract.contract_stop_date.strftime(
#                     "%Y-%m-%d"
#                 ),
#             }

#             return Response(
#                 {
#                     "success": True,
#                     "message": "Drug details retrieved successfully.",
#                     "data": [response_data],
#                 },
#                 status=status.HTTP_200_OK,
#             )

#         except Http404:
#             return Response(
#                 {
#                     "success": False,
#                     "message": "Drug not found.",
#                     "data": None,
#                 },
#                 status=status.HTTP_404_NOT_FOUND,
#             )

#         except Exception as e:
#             return Response(
#                 {
#                     "success": False,
#                     "message": str(e),
#                     "data": None,
#                 },
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )


class DrugFullDetailViewSet(viewsets.ModelViewSet):
    queryset = FSSDrug.objects.all()
    serializer_class = DrugFullDetailSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=["post"])
    def details(self, request, drug_id=None):
        try:
            # Fetch the drug by its primary key (ID)
            drug = get_object_or_404(FSSDrug, id=drug_id)
            
            contract_type = request.data.get('price_type')  # Get contract_type from request body
            price = request.data.get('price')  # Get price from request body
            
            pricing_queryset = FSSPricing.objects.filter(drug=drug)
            if contract_type:
                pricing_queryset = pricing_queryset.filter(price_type=contract_type)
            if price:
                pricing_queryset = pricing_queryset.filter(price=price)

            # Fetch the first pricing that matches the criteria
            general_pricing = pricing_queryset.first()

            # Fetch the first pricing related to the drug
            #pricing = FSSPricing.objects.filter(drug=drug).first()

            # Extract the ingredient and strength of the drug
            ingredient = drug.ingredient
            strength = drug.strength

            # Initialize dictionaries to hold min and max prices for each price type
            price_type_ranges = {
                "FSS": {"min": 0, "max": 0},
                "NC": {"min": 0, "max": 0},
                "Big4": {"min": 0, "max": 0},
            }

            # Query for other drugs with the same ingredient and strength
            related_drugs = FSSDrug.objects.filter(
                ingredient=ingredient, strength=strength
            )

            # Iterate over related drugs to find min and max prices
            for related_drug in related_drugs:
                pricings = FSSPricing.objects.filter(drug=related_drug)
                for pricing in pricings:
                    if pricing.price_type in price_type_ranges:
                        if (
                            price_type_ranges[pricing.price_type]["min"] == 0
                            or pricing.price
                            < price_type_ranges[pricing.price_type]["min"]
                        ):
                            price_type_ranges[pricing.price_type]["min"] = pricing.price
                        if pricing.price > price_type_ranges[pricing.price_type]["max"]:
                            price_type_ranges[pricing.price_type]["max"] = pricing.price

            # Prepare response data from the drug and its related models
            response_data = {
                "drug_id": drug.id,
                "trade_name": drug.trade_name,
                "ingredient": drug.ingredient,
                "strength": drug.strength,
                "package_description": drug.package_description,
                "ndc_code": drug.ndc_with_dashes,
                "contract_number": drug.contract.contract_number,
                "dosage_form": drug.dosage_form,
                "dosage_route": drug.route,
                "notes":drug.notes,
               
                "estimated_annual_quantities": drug.contract.estimated_annual_quantities,  # Fetching from related contract

                "estimated_resolicitation_date":drug.estimated_resolicitation_date,
                "estimated_annual_spend":drug.estimated_annual_spend,
                "offers":drug.offers,
                "manufacture_info": {
                    "manufactured_by": (
                        drug.manufactured_by.name if drug.manufactured_by else None
                    ),
                    "manufactured_for": (
                        drug.manufactured_for.name if drug.manufactured_for else None
                    ),
                    "manufactured_by_address": (
                        drug.manufactured_by.address if drug.manufactured_by else None
                    ),
                    "distributed_by": (
                        drug.distributed_by.name if drug.distributed_by else None
                    ),
                },
                "image_urls": drug.image_urls,
                "price_info": {
                    "price_type": general_pricing.price_type if general_pricing else None,
                    "price": general_pricing.price if general_pricing else None,
                    "price_start_date": (
                        general_pricing.price_start_date.strftime("%Y-%m-%d")
                        if general_pricing
                        else None
                    ),
                    "price_end_date": (
                        general_pricing.price_stop_date.strftime("%Y-%m-%d")
                        if general_pricing
                        else None
                    ),
                    "price_ranges": price_type_ranges,
                },
                "award_info": {
                    "awarded_value": drug.contract.awarded_value,
                    "awarded_quality": drug.contract.estimated_annual_quantities,
                    "awardee": drug.contract.awardee,
                },
            }

            return Response(
                {
                    "success": True,
                    "message": "Drug details retrieved successfully.",
                    "data": response_data,
                },
                status=status.HTTP_200_OK,
            )

        except Http404:
            return Response(
                {
                    "success": False,
                    "message": "No FSS Drug matches the given query.",
                    "data": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        except Exception as e:
            print(f"Error retrieving drug details: {str(e)}")
            return Response(
                {
                    "success": False,
                    "message": str(e),
                    "data": None,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RelatedTradeNamesViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsPagination

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'ingredient',
                openapi.IN_QUERY,
                description="Filter drugs by the ingredient name",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                'dosage_form',
                openapi.IN_QUERY,
                description="Filter drugs by dosage form",
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                'route',
                openapi.IN_QUERY,
                description="Filter drugs by dosage route",
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                "price_type",
                openapi.IN_QUERY,
                description="Filter by price type (FSS, NC, Big4)",
                type=openapi.TYPE_STRING,
                enum=["FSS", "NC", "Big4"],
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search by trade name or NDC code",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "page",
                openapi.IN_QUERY,
                description="Page number for pagination",
                type=openapi.TYPE_INTEGER,
            ),
            openapi.Parameter(
                "page_size",
                openapi.IN_QUERY,
                description="Number of results per page",
                type=openapi.TYPE_INTEGER,
            ),
        ],
        responses={200: RelatedTradeNameSerializer(many=True)},
    )
    @action(detail=False, methods=["get"])
    def related_trade_names(self, request, drug_id=None):
        ingredient = request.query_params.get('ingredient')
        price_type = request.query_params.get('price_type')
        search = request.query_params.get('search')
        dosage_form = request.query_params.get('dosage_form')
        route = request.query_params.get('route')

        # Initialize Q object to accumulate filters
        filters = Q()

        # Apply filters based on provided parameters
        if ingredient:
            filters &= Q(ingredient__icontains=ingredient)
        if dosage_form:
            filters &= Q(dosage_form__icontains=dosage_form)
        if route:
            filters &= Q(route__icontains=route)

        # Apply filters
        drugs = FSSDrug.objects.filter(filters)

        # Filter by price type if provided
        if price_type:
            drugs = drugs.filter(pricings__price_type__iexact=price_type)

        # Apply search filter if provided
        if search:
            drugs = drugs.filter(
                Q(trade_name__icontains=search) |
                Q(ndc_with_dashes__icontains=search)
            )

        # Apply pagination
        paginator = self.pagination_class()
        paginated_drugs = paginator.paginate_queryset(drugs, request)
        serializer = RelatedTradeNameSerializer(paginated_drugs, many=True)

        return paginator.get_paginated_response(serializer.data)


class VendorsInfoViewSet(viewsets.ModelViewSet):
    queryset = FSSVendor.objects.prefetch_related(
        Prefetch(
            "contracts", queryset=FSSContract.objects.all().order_by("contract_number")
        )
    ).order_by("id")
    serializer_class = VendorsInfoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsPagination
    ordering = ["id"]

    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    @action(
        detail=False, methods=["get"]
    )  # Changed to detail=False since we are fetching a list
    def vendors_info(self, request):
        try:
            search_text = request.query_params.get("search", "")
            queryset = self.get_queryset()

            # Apply search filter if search is provided
            if search_text:
                queryset = queryset.filter(
                    Q(vendor_name__icontains=search_text)
                    | Q(contracts__contract_number__icontains=search_text)
                ).distinct()

            vendors = self.paginate_queryset(queryset)

            if vendors is not None:
                vendors_info = [
                    {
                        "id": vendor.id,
                        "vendor_name": vendor.vendor_name,
                        "contracts": [
                            {
                                "contract_number": contract.contract_number,
                                "contract_start_date": contract.contract_start_date.strftime(
                                    "%Y-%m-%d"
                                ),
                                "contract_stop_date": contract.contract_stop_date.strftime(
                                    "%Y-%m-%d"
                                ),
                            }
                            for contract in vendor.contracts.all()
                        ],
                    }
                    for vendor in vendors
                ]
                return self.get_paginated_response(vendors_info)

            return Response(
                {
                    "success": True,
                    "message": "No vendors found.",
                    "data": [],
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        except Exception as e:
            print(f"Error fetching vendors and contracts: {str(e)}")
            return Response(
                {
                    "success": False,
                    "message": f"Error fetching vendors and contracts: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
            
    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search text for filtering across multiple fields",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def export_to_excel(self, request):

        try:
            search_text = request.query_params.get("search", "")
            queryset = self.get_queryset()

            # Apply search filter if search is provided
            if search_text:
                queryset = queryset.filter(
                    Q(vendor_name__icontains=search_text)
                    | Q(contracts__contract_number__icontains=search_text)
                ).distinct()

            vendors = self.paginate_queryset(queryset)
           
            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Vendors Info"

            # Add headers to the worksheet
            headers = [
                "Vendor Name",
                "Contract Number",
                "Contract Start Date",
                "Contract Stop Date",
            ]
            sheet.append(headers)

            # Write data to the worksheet
            for vendor in vendors:
                for contract in vendor.contracts.all():
                    sheet.append(
                        [
                            vendor.vendor_name,
                            contract.contract_number,
                            contract.contract_start_date.strftime("%Y-%m-%d"),
                            contract.contract_stop_date.strftime("%Y-%m-%d"),
                        ]
                    )

            # Set the file name
            file_name = (
                f"vendors_info_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class VendorDetailsViewSet(viewsets.ModelViewSet):
    queryset = FSSVendor.objects.all()
    serializer_class = VendorDetailSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=["get"])
    def vendor_details(self, request, vendor_id=None):
        try:
            # Fetch the vendor by its ID
            vendor = get_object_or_404(FSSVendor, id=vendor_id)

            # Prepare a list of contracts with relevant information
            contracts_info = [
                {
                    "contract_number": contract.contract_number,
                    "contract_start_date": contract.contract_start_date.strftime(
                        "%Y-%m-%d"
                    ),
                    "contract_stop_date": contract.contract_stop_date.strftime(
                        "%Y-%m-%d"
                    ),
                    "drugs": [
                        {
                            "va_class": drug.va_class,
                            "covered": drug.covered,
                            "prime_vendor": drug.prime_vendor,
                        }
                        for drug in contract.drugs.all()
                    ],
                }
                for contract in vendor.contracts.all()
            ]

            response_data = {
                "vendor_name": vendor.vendor_name,
                "notes": vendor.notes,
                "contracts": contracts_info,
            }

            return Response(
                {
                    "success": True,
                    "message": f"Vendor '{vendor.vendor_name}' details retrieved successfully.",
                    "data": response_data,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(f"Error fetching vendor details: {str(e)}")
            return Response(
                {
                    "success": False,
                    "message": f"Error fetching vendor details: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DrugDetailByVendorViewSet(viewsets.ModelViewSet):
    queryset = FSSVendor.objects.all()
    serializer_class = DrugDetailByVendorSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=["get"])
    def drugs_by_vendor(self, request, vendor_id=None):
        try:
            # Fetch the vendor by its ID
            vendor = get_object_or_404(FSSVendor, id=vendor_id)

            # Fetch drugs associated with the vendor
            drugs = FSSDrug.objects.filter(vendor=vendor)

            # Prepare a list of drug details
            drugs_info = [
                {
                    "trade_name": drug.trade_name,
                    "ingredient": drug.ingredient,
                    "strength": drug.strength,
                    "ndc_code": drug.ndc_with_dashes,
                    "price": (
                        FSSPricing.objects.filter(drug=drug).first().price
                        if FSSPricing.objects.filter(drug=drug).first()
                        else None
                    ),
                }
                for drug in drugs
            ]

            return Response(
                {
                    "success": True,
                    "message": f"Drugs associated with vendor '{vendor.vendor_name}' retrieved successfully.",
                    "data": drugs_info,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(f"Error fetching drugs for vendor: {str(e)}")
            return Response(
                {
                    "success": False,
                    "message": f"Error fetching drugs for vendor: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class PotentialLeadViewSet(viewsets.ModelViewSet):
    queryset = PotentialLead.objects.all()
    serializer_class = PotentialLeadSerializer
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'market_status',
                openapi.IN_QUERY,
                description='Filter by market status (e.g., "DISCN" or "RX")',
                type=openapi.TYPE_STRING,
                enum=['DISCN', 'RX'],  # Define the choices here
            ),
        ],
        responses={200: PotentialLeadSerializer(many=True)},
    )
    @action(detail=True, methods=["get"], url_path='retrieve-leads')
    def retrieve_leads(self, request, drug_id=None):
        """
        Retrieve potential leads based on the active ingredient, strength, dosage form, and route
        of a given drug ID with optional filtering by market status.
        """
        # Retrieve drug by ID from the URL path
        if not drug_id:
            return Response(
                {"success": False, "message": "Drug ID is required", "data": None},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            drug = FSSDrug.objects.get(pk=drug_id)
        except FSSDrug.DoesNotExist:
            return Response(
                {"success": False, "message": "Drug not found", "data": None},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Filter potential leads by partial matches on active ingredient, strength, dosage form, and route
        queryset = self.queryset.filter(
            active_ingredient__icontains=drug.ingredient,
            strength__icontains=drug.strength,
            dosage_form__icontains=drug.dosage_form,
            route__icontains=drug.route
        )

        # Get market_status from query parameters and apply the filter
        market_status = request.query_params.get('market_status', None)
        if market_status:
            if market_status not in ['DISCN', 'RX']:  # Validate the input
                return Response(
                    {"success": False, "message": "Invalid market status value.", "data": None},
                    status=status.HTTP_400_BAD_REQUEST
                )
            queryset = queryset.filter(market_status=market_status)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "Potential leads retrieved successfully.",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )
    

# class PotentialLeadViewSet(viewsets.ModelViewSet):
#     queryset = PotentialLead.objects.all()
#     serializer_class = PotentialLeadSerializer
#     permission_classes = [IsAuthenticated]

#     def retrieve(self, request, drug_id=None):
#         """
#         Retrieves potential leads based on the ingredient of a given drug ID.
#         """
#         if not drug_id:
#             return Response(
#                 {"success": False, "message": "Drug ID is required", "data": None},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         try:
#             drug = FSSDrug.objects.get(pk=drug_id)
#         except FSSDrug.DoesNotExist:
#             return Response(
#                 {"success": False, "message": "Drug not found", "data": None},
#                 status=status.HTTP_404_NOT_FOUND,
#             )

#         # Filter potential leads by the active ingredient of the drug
#         potential_leads = self.queryset.filter(
#             active_ingredient=drug.ingredient
#         ).order_by(
#             "-market_status",  # Sort with "RX" first
#             "market_status",  # Then sort by other market_status values
#         )

#         serializer = self.get_serializer(potential_leads, many=True)
#         return Response(
#             {
#                 "success": True,
#                 "message": "Potential leads retrieved successfully based on ingredient",
#                 "data": serializer.data,
#             },
#             status=status.HTTP_200_OK,
#         )


class DrugsByDurationView(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = DashboardResultsPagination
    serializer_class = DrugSerializer

    def get_expiry_date_range(self, months):
        now = timezone.now().date()
        if months == 3:
            start_date = now
            end_date = now + relativedelta(months=3)
        elif months == 6:
            start_date = now + relativedelta(months=3)
            end_date = now + relativedelta(months=6)
        elif months == 9:
            start_date = now + relativedelta(months=6)
            end_date = now + relativedelta(months=9)
        elif months == 12:
            start_date = now + relativedelta(months=9)
            end_date = now + relativedelta(months=12)
        else:
            raise ValueError("Invalid duration. It must be 3, 6, 9, or 12 months.")
        return start_date, end_date

    def get_drugs_by_expiry(self, start_date, end_date, price_type=None):
        contracts = FSSContract.objects.filter(
            contract_stop_date__gt=start_date, contract_stop_date__lte=end_date
        )
        drugs = FSSDrug.objects.filter(contract__in=contracts).order_by("id")

        if price_type:
            drugs = drugs.filter(pricings__price_type=price_type)

        return drugs

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "months",
                openapi.IN_QUERY,
                description="Number of months until contract expiry",
                type=openapi.TYPE_INTEGER,
                enum=[3, 6, 9, 12],
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search term for filtering drugs by NDC, trade name, generic name, vendor name, dosage form, or route",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "price_type",
                openapi.IN_QUERY,
                description="Filter by price type (FSS, NC, Big4)",
                type=openapi.TYPE_STRING,
                enum=["FSS", "NC", "Big4"],
            ),
        ],
        responses={200: DrugSerializer(many=True)},
    )
    @action(detail=False, methods=["get"])
    def list(self, request):
        """
        API endpoint to find Drugs by contract expiry duration.
        """
        months = request.query_params.get("months")
        price_type = request.query_params.get("price_type")
        search = request.query_params.get("search")

        if not months:
            return Response(
                {
                    "success": False,
                    "message": "Please provide a duration in months (3, 6, or 9)",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            months = int(months)
            if months not in [3, 6, 9, 12]:
                raise ValueError
        except ValueError:
            return Response(
                {
                    "success": False,
                    "message": "Duration must be one of the following values: 3, 6, or 9",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date, end_date = self.get_expiry_date_range(months)
        except ValueError as e:
            return Response(
                {"success": False, "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        drugs = self.get_drugs_by_expiry(start_date, end_date, price_type)

        if search:
            drugs = drugs.filter(
                Q(ndc_with_dashes__icontains=search)
                | Q(trade_name__icontains=search)
                | Q(generic_name__icontains=search)
                | Q(vendor__vendor_name__icontains=search)
                | Q(dosage_form__icontains=search)
                | Q(route__icontains=search)
            )
        # Count drugs by price type
        fss_count = drugs.filter(pricings__price_type='FSS').count()
        nc_count = drugs.filter(pricings__price_type='NC').count()
        big4_count = drugs.filter(pricings__price_type='Big4').count()

        total_drugs = FSSDrug.objects.count()
        percentage = (drugs.count() / total_drugs) * 100 if total_drugs > 0 else 0
        serializer = DrugSerializer(
            drugs, many=True, context={"price_type": price_type}
        )

        paginator = self.pagination_class()
        paginated_drugs = paginator.paginate_queryset(drugs, request)
        serializer = DrugSerializer(
            paginated_drugs, many=True, context={"price_type": price_type}
        )

        paginated_response = paginator.get_paginated_response(
            serializer.data, round(percentage)
        )
         # Add counts to the response
        response_data = paginated_response.data
        response_data.update({
            "counts": {
                "FSS": fss_count,
                "NC": nc_count,
                "Big4": big4_count
            }
        })
        return paginated_response

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "months",
                openapi.IN_QUERY,
                description="Number of months until contract expiry",
                type=openapi.TYPE_INTEGER,
                enum=[3, 6, 9, 12],
            ),
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description="Search term for filtering drugs by NDC, trade name, generic name, vendor name, dosage form, or route",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "price_type",
                openapi.IN_QUERY,
                description="Filter by price type (FSS, NC, Big4)",
                type=openapi.TYPE_STRING,
                enum=["FSS", "NC", "Big4"],
            ),
        ],
        responses={
            200: openapi.Response(
                "Excel file",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    @action(detail=False, methods=["get"])
    def export_to_excel(self, request):
        try:
            # Get the expiry date range
            months = request.query_params.get("months")
            price_type = request.query_params.get("price_type")
            search = request.query_params.get("search")

            if not months:
                return Response(
                    {
                        "success": False,
                        "message": "Please provide a duration in months (3, 6, 9, or 12)",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:
                months = int(months)
                if months not in [3, 6, 9, 12]:
                    raise ValueError
            except ValueError:
                return Response(
                    {
                        "success": False,
                        "message": "Duration must be one of the following values: 3, 6, 9, or 12",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            start_date, end_date = self.get_expiry_date_range(months)
            drugs = self.get_drugs_by_expiry(start_date, end_date, price_type)

            # Apply search filters if provided

            if search:
                drugs = drugs.filter(
                    Q(ndc_with_dashes__icontains=search)
                    | Q(trade_name__icontains=search)
                    | Q(generic_name__icontains=search)
                    | Q(vendor__vendor_name__icontains=search)
                    | Q(dosage_form__icontains=search)
                    | Q(route__icontains=search)
                )

            # Serialize the data
            serializer = DrugSerializer(
                drugs, many=True, context={"price_type": price_type}
            )

            # Create a workbook and worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Drugs By Duration Export"

            # Add headers based on the serializer fields
            headers = list(serializer.data[0].keys()) if serializer.data else []
            sheet.append(headers)

            # Write serialized data to the worksheet
            for drug in serializer.data:
                sheet.append([drug.get(header) for header in headers])

            # Set the file name
            file_name = f"drugs_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AccessDrugShortageDataViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AccessDrugShortageData.objects.all().order_by("id")
    permission_classes = [IsAuthenticated]
    serializer_class = AccessDrugShortageDataSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    pagination_class = StandardResultsPagination
    search_fields = ["generic_name", "shortage_status"]  # Fields to search by

    @action(detail=False, methods=["get"])
    def list(self, request, *args, **kwargs):
        """
        Override the default list method to provide a consistent response format.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "All drug shortage data retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def export_to_excel(self, request, *args, **kwargs):
        """
        Export the filtered drug shortage data to an Excel file.
        """
        try:
            # Apply filters and search
            queryset = self.filter_queryset(self.get_queryset())

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Access Drug Shortage Data"

            # Add headers to the worksheet
            headers = ["Generic Name", "Shortage Status"]
            sheet.append(headers)

            # Write data to the worksheet
            for data in queryset:
                sheet.append(
                    [
                        data.generic_name,
                        data.shortage_status,
                    ]
                )

            # Set the file name
            file_name = f"access_drug_shortage_data_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AsphDrugShortageDataViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AsphDrugShortageData.objects.all().order_by("id")
    permission_classes = [IsAuthenticated]
    serializer_class = AsphDrugShortageDataSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    pagination_class = StandardResultsPagination
    search_fields = ["generic_name", "shortage_status"]  # Fields to search by

    @action(detail=False, methods=["get"])
    def list(self, request, *args, **kwargs):
        """
        Override the default list method to provide a consistent response format.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "All drug shortage data retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def export_to_excel(self, request, *args, **kwargs):
        """
        Export the filtered drug shortage data to an Excel file.
        """
        try:
            # Apply filters and search
            queryset = self.filter_queryset(self.get_queryset())

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Asph Drug Shortage Data"

            # Add headers to the worksheet
            headers = [
                "Generic Name",
                "Shortage Status",
                "Created Date",
                "Revision Date",
            ]
            sheet.append(headers)

            # Write data to the worksheet
            for data in queryset:
                sheet.append(
                    [
                        data.generic_name,
                        data.shortage_status,
                        data.created_date,
                        data.revision_date,
                    ]
                )

            # Set the file name
            file_name = f"asph_drug_shortage_data_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class FIOAUniqueNDCViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FOIAUniqueNDCData.objects.all().order_by("id")
    permission_classes = [IsAuthenticated]
    serializer_class = FOIAUniqueNDCDataSerializer
    pagination_class = StandardResultsPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ["ndc_code", "description"]  # Fields to search by
    ordering_fields = [
        "id",
        "ndc_code",
        "total_quantity_purchased",
        "total_publishable_dollars_spent",
    ]  # Fields to sort by
    ordering = ["id"]  # Default ordering

    @action(detail=False, methods=["get"])
    def list(self, request, *args, **kwargs):
        """
        Override the default list method to provide a consistent response format.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "All purchase records retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["get"])
    def retrieve_by_id(self, request, pk=None):
        """
        Custom action to retrieve FOIAUniqueNDCData by id with a consistent response format.
        """
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(
                {
                    "success": True,
                    "message": "Data retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
        except FOIAUniqueNDCData.DoesNotExist:
            return Response(
                {"success": False, "message": "Data not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=False, methods=["get"])
    def export_to_excel(self, request, *args, **kwargs):
        """
        Export the filtered FOIA drugs data to an Excel file.
        """
        try:
            # Apply filters, search, and ordering
            queryset = self.filter_queryset(self.get_queryset())

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "FOIA Drugs Data"

            headers = [
                "NDC Code",
                "Ingredient Name",
                "Quantity Purchased",
                "Dollars Spent",
                "Description",
            ]
            sheet.append(headers)

            # Write data to the worksheet
            for data in queryset:
                sheet.append(
                    [
                        data.ndc_code if data.ndc_code else "",
                        data.ingredient,
                        data.total_quantity_purchased,
                        data.total_publishable_dollars_spent,
                        data.description,
                    ]
                )

            # Set the file name
            file_name = f"foia_drugs_data_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class FIOADrugsDataByNdcViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = (
        FOIADrugsData.objects.select_related("station_data", "ndc_code")
        .all()
        .order_by("id")
    )
    serializer_class = FOIADrugsDataSerializer
    pagination_class = StandardResultsPagination
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        method="get",
        manual_parameters=[
            openapi.Parameter(
                "ndc_code",
                openapi.IN_QUERY,
                description="Search text for ndc_code",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    @action(detail=False, methods=["get"])
    def find_by_ndc_code(self, request, *args, **kwargs):
        """
        Custom action to find FOIADrugsData records by ndc_code.
        """
        ndc_code = request.query_params.get("ndc_code")
        if not ndc_code:
            return Response(
                {"success": False, "message": "Please provide an ndc_code"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        drugs_data = FOIADrugsData.objects.filter(ndc_code__ndc_code=ndc_code)
        page = self.paginate_queryset(drugs_data)
        if page is not None:
            serializer = FOIADrugsDataSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = FOIADrugsDataSerializer(drugs_data, many=True)
        return Response(
            {
                "success": True,
                "message": f"Data for NDC code {ndc_code} retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )


class FIOADrugsDataViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FOIADrugsData.objects.select_related("ndc_code").all().order_by("id")
    serializer_class = FOIADrugsDataSerializer
    pagination_class = StandardResultsPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = [
        "ndc_code__ndc_code",
        "mckesson_station_number",
    ]  # Fields to search by
    ordering_fields = [
        "id",
        "ndc_code",
        "quantity_purchased",
        "publishable_dollars_spent",
    ]  # Fields to sort by
    ordering = ["id"]  # Default ordering
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def list(self, request, *args, **kwargs):
        """
        Override the default list method to provide a consistent response format.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "All FOIA drug data retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )


class DODDrugDataViewSet(viewsets.ModelViewSet):

    queryset = DODDrugData.objects.all().order_by("id")
    serializer_class = DODDrugsDataSerializer
    pagination_class = StandardResultsPagination
    filter_backends = (DjangoFilterBackend, SearchFilter)
    search_fields = ["ndc_code", "description"]
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def list(self, request, *args, **kwargs):
        """
        Override the default list method to provide a consistent response format.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "success": True,
                "message": "All drug data retrieved successfully",
                "data": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def export_to_excel(self, request, *args, **kwargs):
        """
        Export the filtered DOD drug data to an Excel file.
        """
        try:
            # Apply filters and search
            queryset = self.filter_queryset(self.get_queryset())

            # Create a workbook and a worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "DOD Drugs Data"

            # Add headers to the worksheet
            headers = ["NDC Code", "Description", "Quantity", "Price"]
            sheet.append(headers)

            # Write data to the worksheet
            for data in queryset:
                sheet.append(
                    [
                        data.ndc_code,
                        data.description,
                        data.quantity,
                        data.price,
                    ]
                )

            # Set the file name
            file_name = (
                f"dod_drugs_data_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            # Create a HTTP response with an Excel content type
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename={file_name}"

            # Save the workbook to the response
            workbook.save(response)

            return response

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error exporting data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DashboardDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get_expiry_counts(self):
        now = timezone.now().date()
        expiry_counts = {}

        # Calculate expiring contracts for 3, 6, 9, and 12 months
        periods = [3, 6, 9, 12]
        for months in periods:
            start_date = now if months == 3 else now + relativedelta(months=months - 3)
            end_date = now + relativedelta(months=months)
            contracts = FSSContract.objects.filter(
                contract_stop_date__gt=start_date, contract_stop_date__lte=end_date
            )
            
            # Filter drugs related to these expiring contracts
            drugs = FSSDrug.objects.filter(contract__in=contracts)
            
            # Count the number of drugs by price type within the expiring contracts
            fss_count = drugs.filter(pricings__price_type='FSS').count()
            nc_count = drugs.filter(pricings__price_type='NC').count()
            big4_count = drugs.filter(pricings__price_type='Big4').count()
            
            # Calculate the total number of expiring drugs for this period
            total_count = fss_count + nc_count + big4_count
            
            # Store the counts in a dictionary
            expiry_counts[f"{months}_months"] = {
                "total": total_count,
                "FSS": fss_count,
                "NC": nc_count,
                "Big4": big4_count
            }

        return expiry_counts


    def get_unique_ndc_counts(self):
        fss_count = FSSDrug.objects.values("ndc_with_dashes").distinct().count()
        foia_count = FOIAUniqueNDCData.objects.values("ndc_code").distinct().count()
        dod_count = DODDrugData.objects.values("ndc_code").distinct().count()
        return {"FSS": fss_count, "FOIA": foia_count, "DOD": dod_count}
    
    def get_price_type_counts(self):
        drugs = FSSDrug.objects.all()
        fss_count = drugs.filter(pricings__price_type='FSS').count()
        nc_count = drugs.filter(pricings__price_type='NC').count()
        big4_count = drugs.filter(pricings__price_type='Big4').count()
        return {"FSS": fss_count, "NC": nc_count, "Big4": big4_count}
    def get_last_updates(self):
        last_updates = {}
        drug_types = ["FSS", "FOIA", "DOD"]
        for drug_type in drug_types:
            record = (
                DataInsertionRecord.objects.filter(drug_type=drug_type)
                .order_by("-date_inserted")
                .first()
            )
            last_updates[drug_type] = (
                record.date_inserted.strftime("%Y-%m-%d %H:%M:%S") if record else None
            )
        return last_updates

    def get(self, request):
        """
        API endpoint to get dashboard data.
        """
        try:
            expiry_counts = self.get_expiry_counts()
            unique_ndc_counts = self.get_unique_ndc_counts()
            last_updates = self.get_last_updates()
            price_type_counts = self.get_price_type_counts()  # Added this line

            dashboard_data = {
                "expiring_contracts": expiry_counts,
                "unique_ndcs": unique_ndc_counts,
                "price_type_counts": price_type_counts,  # Added this line
                "last_updates": last_updates,
            }

            return Response(
                {
                    "success": True,
                    "message": "Data retrieved successfully",
                    "data": dashboard_data,
                },
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            return Response(
                {"success": False, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class LatestScrapingStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        # Retrieve the latest scraping task status from the database
        latest_scraping_task = ScrapingStatus.objects.order_by("-start_time").first()

        if latest_scraping_task:
            serializer = ScrapingStatusSerializer(latest_scraping_task)
            return Response(
                {
                    "success": True,
                    "message": "Latest Scraping data retrieved successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )
        else:
            message = "No scraping task found."
            return Response(
                {"success": False, "message": message}, status=status.HTTP_200_OK
            )


class FOIAMonthlyStatsListView(viewsets.ModelViewSet):
    pagination_class = StandardResultsPagination
    queryset = FOIAMonthlyStats.objects.all().order_by("id")
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter('ndc', openapi.IN_QUERY, description="Search by NDC", type=openapi.TYPE_STRING),
            openapi.Parameter('month', openapi.IN_QUERY, description="Search by Month", type=openapi.TYPE_STRING),
            openapi.Parameter('year', openapi.IN_QUERY, description="Search by Year", type=openapi.TYPE_INTEGER),
            openapi.Parameter('product_name', openapi.IN_QUERY, description="Search by Product Name", type=openapi.TYPE_STRING),
            openapi.Parameter('search', openapi.IN_QUERY, description="Generic search across multiple fields", type=openapi.TYPE_STRING),
            openapi.Parameter('start_date', openapi.IN_QUERY, description="Start Date (YYYY-MM)", type=openapi.TYPE_STRING),
            openapi.Parameter('end_date', openapi.IN_QUERY, description="End Date (YYYY-MM)", type=openapi.TYPE_STRING),
        ],
        responses={200: FOIAMonthlyStatsSerializer(many=True)}, 
        operation_description="Search FOIA Monthly Stats by NDC, Month, Year, Product Name, or Generic Search, including a Date Range filter"
    )
    def get(self, request, *args, **kwargs):
        ndc = request.query_params.get('ndc', '')
        month = request.query_params.get('month', '')
        year = request.query_params.get('year', '')
        product_name = request.query_params.get('product_name', '')
        search = request.query_params.get('search', '')

        # Date range filters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        queryset = FOIAMonthlyStats.objects.all()

        # Apply individual filters
        if ndc:
            queryset = queryset.filter(Q(ndc__icontains=ndc))
        if month:
            queryset = queryset.filter(Q(month__icontains=month))
        if year:
            queryset = queryset.filter(Q(year__icontains=year))
        if product_name:
            queryset = queryset.filter(Q(product_name__icontains=product_name))

        if search:
            queryset = queryset.filter(
                Q(ndc__icontains=search) |
                Q(product_name__icontains=search) |
                Q(strength__icontains=search) |
                Q(month__icontains=search) |
                Q(year__icontains=search)
            )

        # Date range filter
        if start_date and end_date:
            try:
                # Convert start_date and end_date to strings for comparison
                start_date_str = datetime.strptime(start_date, "%Y-%m").strftime("%Y-%m")
                end_date_str = datetime.strptime(end_date, "%Y-%m").strftime("%Y-%m")

                # Ensure start_date is earlier than or equal to end_date
                if start_date_str > end_date_str:
                    return Response(
                        {"success": False, "message": "start_date should be earlier than end_date."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Annotate queryset with numeric month using Case-When statements
                queryset = queryset.annotate(
                    numeric_month=Case(
                        When(month="January", then=Value("01")),
                        When(month="February", then=Value("02")),
                        When(month="March", then=Value("03")),
                        When(month="April", then=Value("04")),
                        When(month="May", then=Value("05")),
                        When(month="June", then=Value("06")),
                        When(month="July", then=Value("07")),
                        When(month="August", then=Value("08")),
                        When(month="September", then=Value("09")),
                        When(month="October", then=Value("10")),
                        When(month="November", then=Value("11")),
                        When(month="December", then=Value("12")),
                        output_field=CharField()
                    )
                )

                # Concatenate year and numeric month to form year_month in "YYYY-MM" format
                queryset = queryset.annotate(
                    year_month=Concat('year', Value('-'), 'numeric_month', output_field=CharField())
                )

                # Apply range filter on year_month
                queryset = queryset.filter(
                    year_month__gte=start_date_str,
                    year_month__lte=end_date_str
                ).order_by("year_month")

            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid date format. Use YYYY-MM for start_date and end_date."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if not queryset.exists():
            return Response(
                {
                    "success": False,
                    "message": "No records found for the given query parameters",
                    "data": [],
                },
                status=status.HTTP_404_NOT_FOUND
            )

        paginator = self.pagination_class()
        paginated_queryset = paginator.paginate_queryset(queryset, request)
        serializer = FOIAMonthlyStatsSerializer(paginated_queryset, many=True)
        
        return paginator.get_paginated_response(serializer.data)

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter('ndc', openapi.IN_QUERY, description="Search by NDC", type=openapi.TYPE_STRING),
            openapi.Parameter('month', openapi.IN_QUERY, description="Search by Month", type=openapi.TYPE_STRING),
            openapi.Parameter('year', openapi.IN_QUERY, description="Search by Year", type=openapi.TYPE_INTEGER),
            openapi.Parameter('product_name', openapi.IN_QUERY, description="Search by Product Name", type=openapi.TYPE_STRING),
            openapi.Parameter('search', openapi.IN_QUERY, description="Generic search across multiple fields", type=openapi.TYPE_STRING),
            openapi.Parameter('start_date', openapi.IN_QUERY, description="Start Date (YYYY-MM)", type=openapi.TYPE_STRING),
            openapi.Parameter('end_date', openapi.IN_QUERY, description="End Date (YYYY-MM)", type=openapi.TYPE_STRING),
        ],
        responses={200: FOIAMonthlyStatsSerializer(many=True)}, 
        operation_description="Search FOIA Monthly Stats by NDC, Month, Year, Product Name, or Generic Search"
    )
    def export_to_excel(self, request, *args, **kwargs):
        # Retrieve query parameters
        ndc = request.query_params.get('ndc', '')
        month = request.query_params.get('month', '')
        year = request.query_params.get('year', '')
        product_name = request.query_params.get('product_name', '')
        search = request.query_params.get('search', '')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Build the queryset based on search parameters
        queryset = FOIAMonthlyStats.objects.all()

        # Apply exact match filters
        if ndc:
            queryset = queryset.filter(Q(ndc__exact=ndc))  # Exact match for NDC
        if month:
            queryset = queryset.filter(Q(month__exact=month))  # Exact match for month
        if year:
            queryset = queryset.filter(Q(year__exact=year))  # Exact match for year
        if product_name:
            queryset = queryset.filter(Q(product_name__icontains=product_name))  # Partial match for product_name

        # Apply generic search filter across multiple fields
        if search:
            queryset = queryset.filter(
                Q(ndc__icontains=search) |
                Q(product_name__icontains=search) |
                Q(strength__icontains=search) |
                Q(month__icontains=search) |
                Q(year__icontains=search)
            )

        # Apply date range filtering if start_date and end_date are provided
        if start_date and end_date:
            try:
                # Convert start_date and end_date to "YYYY-MM" format
                start_date_str = datetime.strptime(start_date, "%Y-%m").strftime("%Y-%m")
                end_date_str = datetime.strptime(end_date, "%Y-%m").strftime("%Y-%m")

                # Ensure start_date is earlier than or equal to end_date
                if start_date_str > end_date_str:
                    return Response(
                        {"success": False, "message": "start_date should be earlier than end_date."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Annotate with numeric month and construct year_month for range filtering
                queryset = queryset.annotate(
                    numeric_month=Case(
                        When(month="January", then=Value("01")),
                        When(month="February", then=Value("02")),
                        When(month="March", then=Value("03")),
                        When(month="April", then=Value("04")),
                        When(month="May", then=Value("05")),
                        When(month="June", then=Value("06")),
                        When(month="July", then=Value("07")),
                        When(month="August", then=Value("08")),
                        When(month="September", then=Value("09")),
                        When(month="October", then=Value("10")),
                        When(month="November", then=Value("11")),
                        When(month="December", then=Value("12")),
                        output_field=CharField()
                    )
                ).annotate(
                    year_month=Concat('year', Value('-'), 'numeric_month', output_field=CharField())
                ).filter(
                    year_month__gte=start_date_str, year_month__lte=end_date_str
                ).order_by("year_month")

            except ValueError:
                return Response(
                    {"success": False, "message": "Invalid date format. Use YYYY-MM for start_date and end_date."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Apply ordering to ensure consistency for Excel export
        if not (start_date and end_date):
            queryset = queryset.order_by("id")

        # If no records found, return a 404 response
        if not queryset.exists():
            return Response(
                {
                    "success": False,
                    "message": "No records found for the given query parameters",
                    "data": [],
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "FOIA Monthly Stats"

        # Define the headers for the Excel file
        headers = ["NDC", "Product Name", "Strength", "Total Dollar Spent", "Total Units Purchased", 
                   "Min Purchase Price", "Max Purchase Price", "Month", "Year"]
        worksheet.append(headers)

        # Add rows from the queryset to the Excel file
        for record in queryset:
            row = [
                record.ndc,
                record.product_name,
                record.strength,
                record.total_dollar_spent,
                record.total_units_purchased,
                record.min_purchase_price,
                record.max_purchase_price,
                record.month,
                record.year
            ]
            worksheet.append(row)

        # Save the Excel file to a temporary in-memory buffer
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=foia_monthly_stats.xlsx'
        workbook.save(response)  # Write the content to the response object

        return response